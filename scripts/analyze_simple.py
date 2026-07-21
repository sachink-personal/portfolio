#!/usr/bin/env python3
"""Simple file analyzer without pandas."""
import sys
import os
import csv

os.chdir('d:/delete/portfolio_g')

print("=" * 80)
print("CHARTINK CSV ANALYSIS")
print("=" * 80)

try:
    with open('downloads/for_Portfolio_7_13_2026.csv', 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
        
        print(f"\nTotal rows: {len(rows)}")
        print(f"\nColumns ({len(rows[0])}):")
        for i, col in enumerate(rows[0], 1):
            print(f"  {i}. {col}")
        
        print(f"\nFirst 2 data rows:")
        for row_idx in range(1, min(3, len(rows))):
            print(f"\n  Row {row_idx}:")
            for col_idx, (col_name, val) in enumerate(zip(rows[0], rows[row_idx]), 1):
                print(f"    {col_name}: {val}")
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("SCREENER EXCEL ANALYSIS (openpyxl)")
print("=" * 80)

try:
    from openpyxl import load_workbook
    
    wb = load_workbook('downloads/screener.xlsx')
    ws = wb.active
    
    # Get all data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    if data:
        headers = data[0]
        print(f"\nShape: {len(data)} rows")
        print(f"\nColumns ({len(headers)}):")
        for i, col in enumerate(headers, 1):
            print(f"  {i}. {col}")
        
        print(f"\nFirst 2 data rows:")
        for row_idx in range(1, min(3, len(data))):
            print(f"\n  Row {row_idx}:")
            for col_idx, (col_name, val) in enumerate(zip(headers, data[row_idx]), 1):
                print(f"    {col_name}: {val}")

except ImportError:
    print("openpyxl not installed, trying pandas...")
    try:
        import pandas as pd
        df = pd.read_excel('downloads/screener.xlsx')
        print(f"\nShape: {df.shape}")
        print(f"\nColumns:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")
        print(f"\nFirst 2 rows:\n{df.head(2)}")
    except Exception as e:
        print(f"Error: {e}")
